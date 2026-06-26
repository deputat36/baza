"""
Генератор XLSX для базы знаний.

Назначение:
- собрать стартовую Excel-книгу из CSV-файлов репозитория;
- использовать её как промежуточный вариант перед переносом в закрытую Google Таблицу;
- не хранить в файле реальные внутренние контакты и приватные данные.

Запуск из корня репозитория:
python scripts/tools/build_workbook_from_csv.py

Результат:
build/baza-knowledge-mvp.xlsx
"""

import csv
from pathlib import Path

from workbook_config import BUILD_DIR, OUTPUT_FILE, ROOT, SHEETS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Не установлен openpyxl. Установите командой: pip install openpyxl") from exc


def read_csv_rows(path: Path):
    if not path.exists():
        return []

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.reader(file))


def append_csv_files(sheet, files):
    current_row = 1

    for relative_path in files:
        path = ROOT / relative_path
        rows = read_csv_rows(path)
        if not rows:
            continue

        sheet.cell(row=current_row, column=1, value=f"Источник: {relative_path}")
        sheet.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        for row in rows:
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        current_row += 1


def format_sheet(sheet):
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    source_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    for row in sheet.iter_rows():
        first_cell_value = row[0].value if row else None

        if isinstance(first_cell_value, str) and first_cell_value.startswith("Источник:"):
            for cell in row:
                cell.fill = source_fill
                cell.font = Font(bold=True)

        if row[0].row > 1 and row[0].row <= sheet.max_row:
            previous_value = sheet.cell(row=row[0].row - 1, column=1).value
            if isinstance(previous_value, str) and previous_value.startswith("Источник:"):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(wrap_text=True)

    for col_idx in range(1, min(sheet.max_column, 20) + 1):
        max_length = 10
        column_letter = get_column_letter(col_idx)

        for cell in sheet[column_letter]:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, min(len(value), 45))

        sheet.column_dimensions[column_letter].width = max_length + 2


def create_main_sheet(workbook):
    sheet = workbook.active
    sheet.title = "Главная"

    rows = [
        ["База знаний специалиста по недвижимости", "Стартовая XLSX-сборка"],
        ["Назначение", "Промежуточный файл для проверки структуры перед Google Таблицей"],
        ["Безопасность", "Не добавлять реальные внутренние контакты и персональные данные"],
        ["Следующий шаг", "Импортировать данные в закрытую Google Таблицу"],
        ["Версия", "v0.1 / подготовка к v0.2"],
    ]

    for row in rows:
        sheet.append(row)

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 80

    for cell in sheet[1]:
        cell.font = Font(bold=True)


def build_workbook():
    BUILD_DIR.mkdir(exist_ok=True)

    workbook = Workbook()
    create_main_sheet(workbook)

    for sheet_name, files in SHEETS:
        sheet = workbook.create_sheet(sheet_name)
        append_csv_files(sheet, files)
        format_sheet(sheet)

    workbook.save(OUTPUT_FILE)
    print(f"Готово: {OUTPUT_FILE}")


if __name__ == "__main__":
    build_workbook()
